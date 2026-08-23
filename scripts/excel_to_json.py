import sys
import json
import unicodedata
import openpyxl

def normalize(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    return s

def main():
    if len(sys.argv) < 3:
        print("Uso: python excel_to_json.py <input.xlsx> <output.json>", file=sys.stderr)
        sys.exit(1)
    input_path = sys.argv[1]
    output_path = sys.argv[2]

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]

    # Encontrar la fila de encabezados (busca "codigo" y "producto")
    header_row_idx = None
    headers = {}
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        norm_vals = [normalize(v) for v in row_vals]
        if 'codigo' in norm_vals and 'producto' in norm_vals:
            header_row_idx = row_idx
            for col_idx, val in enumerate(norm_vals, start=1):
                headers[val] = col_idx
            break

    if header_row_idx is None:
        print("No se encontro la fila de encabezados (codigo/producto)", file=sys.stderr)
        sys.exit(1)

    def col(name_options):
        for name in name_options:
            if name in headers:
                return headers[name]
        return None

    c_codigo = col(['codigo'])
    c_producto = col(['producto'])
    c_categoria = col(['categoria'])
    c_subrubro = col(['subrubro'])
    c_existencia = col(['exisistencia (bulto)', 'existencia (bulto)', 'existencia'])
    c_pack = col(['pack'])
    c_costo = col(['costo unitario'])
    c_pu = col(['precio unitario'])
    c_pb = col(['precio bulto'])
    c_pe = col(['precio especial'])
    c_activo = col(['activo'])
    c_especial = col(['especial'])

    productos = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        nombre = ws.cell(row=row_idx, column=c_producto).value if c_producto else None
        if not nombre:
            continue
        def g(c):
            return ws.cell(row=row_idx, column=c).value if c else None

        especial_val = normalize(g(c_especial))
        productos.append({
            'codigo': str(g(c_codigo)) if g(c_codigo) is not None else None,
            'nombre': str(nombre).strip(),
            'categoria': str(g(c_categoria) or '').strip(),
            'subrubro': str(g(c_subrubro) or '').strip(),
            'existencia_bulto': float(g(c_existencia) or 0),
            'pack': int(g(c_pack) or 1),
            'costo_unitario': float(g(c_costo) or 0),
            'precio_unitario': float(g(c_pu) or 0),
            'precio_bulto': float(g(c_pb) or 0),
            'precio_especial': float(g(c_pe) or 0),
            'activo': bool(g(c_activo)),
            'permite_ajuste_precio': especial_val == 'si',
        })

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(productos, f, ensure_ascii=False, indent=2)

    print(f"OK: {len(productos)} productos exportados a {output_path}")

if __name__ == '__main__':
    main()
